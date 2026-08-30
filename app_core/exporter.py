from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def export_missing(rows,path):
    wb=Workbook(); ws=wb.active; ws.title='Faltando'
    headers=['Código de Barras','Descrição','Quantidade Faltando','Localização SS']; ws.append(headers)
    for c in ws[1]: c.font=Font(bold=True); c.alignment=Alignment(horizontal='center',vertical='center')
    for r in rows:
        missing=max(int(r['quantity_expected'])-int(r['quantity_checked']),0)
        if missing<=0: continue
        ws.append([r['barcode'] or r['sku'],r['description'],missing,r['location_simpleset']])
    widths=[20,55,20,24]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for row in ws.iter_rows(min_row=2):
        row[1].alignment=Alignment(wrap_text=True,vertical='top')
        row[0].alignment=Alignment(horizontal='center'); row[2].alignment=Alignment(horizontal='center'); row[3].alignment=Alignment(horizontal='center')
        ws.row_dimensions[row[0].row].height=36
    wb.save(path); return path
