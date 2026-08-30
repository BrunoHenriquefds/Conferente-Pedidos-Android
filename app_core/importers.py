from __future__ import annotations
from pathlib import Path
import re, unicodedata
from openpyxl import load_workbook


def clean(v):
    if v is None: return ''
    s=str(v).strip()
    return '' if s.lower() in {'nan','none'} else s

def number(v, default=0.0):
    if v is None or v=='': return default
    if isinstance(v,(int,float)): return float(v)
    s=str(v).strip()
    if ',' in s: s=s.replace('.','').replace(',','.')
    try: return float(s)
    except: return default

def norm_name(v):
    s=unicodedata.normalize('NFKD', str(v or '')).encode('ascii','ignore').decode('ascii')
    return re.sub(r'\s+',' ',s.strip().casefold())

def normalize_document(v): return re.sub(r'\D','',clean(v))
def normalize_code(v): return re.sub(r'\.0$','',clean(v)).strip()

def normalize_tracking(v):
    s=re.sub(r'\s+','',clean(v)).upper()
    # MEL47640886859FMXDF01 -> 47640886859
    m=re.search(r'(\d{8,})',s)
    return m.group(1) if m else s

def _sheet_rows(path):
    wb=load_workbook(Path(path), read_only=True, data_only=True)
    ws=wb.active
    rows=list(ws.iter_rows(values_only=True))
    return rows

def _detect_header(rows):
    targets={norm_name(x) for x in ('N.º de venda','Nº de venda','Nome Cliente','Produto','Título do anúncio')}
    for i,row in enumerate(rows[:30]):
        vals={norm_name(x) for x in row if x is not None}
        if vals & targets: return i
    return 0

def _records(path):
    rows=_sheet_rows(path); hi=_detect_header(rows)
    raw_headers=[clean(x) for x in rows[hi]]
    headers=[]; seen={}
    for h in raw_headers:
        if not h:
            headers.append(''); continue
        key=norm_name(h); seen[key]=seen.get(key,0)+1
        headers.append(h if seen[key]==1 else f'{h}__{seen[key]}')
    out=[]
    for r in rows[hi+1:]:
        d={headers[i]: r[i] if i<len(r) else None for i in range(len(headers)) if headers[i]}
        out.append(d)
    return headers,out

def _cmap(headers):
    cm={}
    for h in headers:
        if not h: continue
        base=h.split('__',1)[0]
        cm.setdefault(norm_name(base),h)
    return cm

def _get(row,cmap,*names):
    for n in names:
        k=cmap.get(norm_name(n))
        if k:
            v=clean(row.get(k))
            if v: return v
    return ''

def import_excel(path):
    headers, recs=_records(path); cm=_cmap(headers)
    if norm_name('N.º de venda') in cm or norm_name('Título do anúncio') in cm:
        return import_mercado_livre(headers,recs,Path(path).name)
    if norm_name('Nome Cliente') in cm and norm_name('Produto') in cm:
        return import_pedidos(headers,recs,Path(path).name)
    raise ValueError('Formato não reconhecido: colunas esperadas não encontradas.')

def import_mercado_livre(headers,recs,source):
    cm=_cmap(headers); out=[]
    for r in recs:
        order=_get(r,cm,'N.º de venda','Nº de venda','Numero da venda')
        if not order: continue
        sku=normalize_code(_get(r,cm,'SKU','Código','Codigo'))
        if not sku: continue
        tracking=_get(r,cm,'Número de rastreamento','Numero de rastreamento')
        out.append(dict(
            order_id=order,buyer_name=_get(r,cm,'Comprador','Dados pessoais ou da empresa'),
            buyer_document=normalize_document(_get(r,cm,'CPF','Tipo e número do documento')),
            platform=_get(r,cm,'Canal de venda') or 'Mercado Livre', sku=sku, barcode=sku,
            description=_get(r,cm,'Título do anúncio','Descricao'),
            quantity_expected=int(number(_get(r,cm,'Unidades'),1)),
            unit_price=number(_get(r,cm,'Preço unitário de venda do anúncio (BRL)'),0),
            location_nerus='', location_simpleset='', source=source,
            tracking_original=tracking, tracking_normalized=normalize_tracking(tracking)
        ))
    return out

def import_pedidos(headers,recs,source):
    cm=_cmap(headers); out=[]
    for r in recs:
        code=normalize_code(_get(r,cm,'Produto','EAN','Código de barras','Codigo de barras','SKU'))
        if not code: continue
        out.append(dict(order_id=_get(r,cm,'Pedido Cliente','Pedido ERP','Pedido SS','Pedido Original'),
            buyer_name=_get(r,cm,'Nome Cliente'), buyer_document=normalize_document(_get(r,cm,'CPF/CNPJ')),
            platform=_get(r,cm,'Canal','Origem'), sku=code, barcode=code,
            description=_get(r,cm,'Descrição','Descricao'), quantity_expected=int(number(_get(r,cm,'Quantidade'),1)),
            unit_price=number(_get(r,cm,'Preço','Preco'),0), location_nerus=_get(r,cm,'Localização'),
            location_simpleset=_get(r,cm,'Localização SS','Localizacao Simpleset'), source=source,
            tracking_original=_get(r,cm,'Número de rastreamento','Numero de rastreamento'),
            tracking_normalized=normalize_tracking(_get(r,cm,'Número de rastreamento','Numero de rastreamento'))))
    return out

def import_locations_excel(path):
    headers,recs=_records(path); cm=_cmap(headers)
    code_names=('ISBN','EAN','Cod_Barras','Código de barras','Codigo de barras','Produto','SKU','Código','Codigo')
    loc_names=('Localização SS','Localizacao SS','Localização Simpleset','Localizacao Simpleset')
    code_col=next((cm.get(norm_name(x)) for x in code_names if cm.get(norm_name(x))),None)
    loc_col=next((cm.get(norm_name(x)) for x in loc_names if cm.get(norm_name(x))),None)
    if not code_col or not loc_col: raise ValueError('Planilha precisa de código e Localização SS.')
    by={}
    for r in recs:
        c=normalize_code(r.get(code_col)); l=clean(r.get(loc_col))
        if c and l: by[c]=l
    if not by: raise ValueError('Nenhuma localização preenchida encontrada.')
    return [{'barcode':c,'location_simpleset':l} for c,l in by.items()]

def import_negative_excel(path):
    headers,recs=_records(path); cm=_cmap(headers); out=[]
    for r in recs:
        code=normalize_code(_get(r,cm,'Cod_Barras','Código de barras','Codigo de barras','EAN'))
        if not code: continue
        qty=int(number(_get(r,cm,'Qtty','Quantidade'),0))
        if qty>=0: continue
        out.append(dict(barcode=code,negative_quantity=qty,product=_get(r,cm,'Produto'),
            grade=_get(r,cm,'grade','Grade'),last_purchase=_get(r,cm,'Data_Ultima_Compra','Data Ultima Compra'),
            profit_center=_get(r,cm,'Centro_lucro','Centro de lucro')))
    if not out: raise ValueError('Nenhum produto com quantidade negativa encontrado.')
    return out
